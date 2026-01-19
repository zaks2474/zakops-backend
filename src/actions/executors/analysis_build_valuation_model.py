from __future__ import annotations

from typing import Any, Dict, Optional

from actions.engine.models import ActionError, ActionPayload, ArtifactMetadata, now_utc_iso
from actions.executors._artifacts import resolve_action_artifact_dir
from actions.executors.base import ActionExecutionError, ActionExecutor, ExecutionContext, ExecutionResult


class BuildValuationModelExecutor(ActionExecutor):
    action_type = "ANALYSIS.BUILD_VALUATION_MODEL"

    def validate(self, payload: ActionPayload) -> tuple[bool, Optional[str]]:
        return True, None

    def execute(self, payload: ActionPayload, ctx: ExecutionContext) -> ExecutionResult:
        inputs = payload.inputs or {}
        revenue = inputs.get("revenue")
        ebitda = inputs.get("ebitda")
        multiple = float(inputs.get("multiple") or 5.0)
        notes = str(inputs.get("notes") or "").strip()

        try:
            from openpyxl import Workbook  # type: ignore
        except Exception as e:
            raise ActionExecutionError(
                ActionError(
                    code="openpyxl_missing",
                    message=f"openpyxl is required for XLSX generation: {type(e).__name__}",
                    category="dependency",
                    retryable=False,
                )
            )

        wb = Workbook()
        ws = wb.active
        ws.title = "Valuation"

        ws["A1"] = "Revenue"
        ws["B1"] = revenue if revenue is not None else ""
        ws["A2"] = "EBITDA"
        ws["B2"] = ebitda if ebitda is not None else ""
        ws["A3"] = "Multiple"
        ws["B3"] = multiple
        ws["A5"] = "Enterprise Value"
        ws["B5"] = "=B2*B3"
        ws["A7"] = "Notes"
        ws["B7"] = notes

        out_dir = resolve_action_artifact_dir(ctx)
        xlsx_path = out_dir / "valuation_model.xlsx"
        wb.save(str(xlsx_path))

        artifacts = [
            ArtifactMetadata(
                filename=xlsx_path.name,
                mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                path=str(xlsx_path),
                created_at=now_utc_iso(),
            )
        ]
        outputs: Dict[str, Any] = {"multiple": multiple}
        return ExecutionResult(outputs=outputs, artifacts=artifacts)

